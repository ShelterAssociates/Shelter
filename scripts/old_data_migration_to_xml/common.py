#	Script written on Python 3.6
# 	Author : Parag Fulzele
#	Description : Common methods used in convert shelter old database data into XML files script
#

import os
from time import gmtime, strftime
import xlrd
import openpyxl
import psycopg2
import datetime
import dicttoxml
import xml.etree.ElementTree as ET
import requests
import shutil
import traceback

requests.packages.urllib3.disable_warnings()

from local_settings import *


# root path for files to read
root_folder_path = os.path.dirname(os.path.abspath(__file__))

# log folder path 
log_folder_path = os.path.join(root_folder_path, 'xml_output', 'log')

# dictionary use to set mapping for survey after reading xls and xlsx
question_map_dict = {}
question_option_map_dict = {}

option_dict = {}
city_ward_slum_dict = {}

# dictionary use to set option/parameter use by method
options_dict = {
	'project': None,
		
	'survey': None,
	'survey2': None,
	
	'mapped_excelFile': None,
	
	# xml values to be set while generating xml
	'xml_root': '',
	'xml_root_attr_id': '',
	'xml_root_attr_version': '',
	'formhub_uuid': '',
	
	'output_path': None,
	
	'log_folder_path': None,
}

# get list of all slum and slum code
qry_slum_list = "select distinct slum_id, slum_code from ray_survey_slumsurveymetadata where survey_id = %s order by slum_id"

# process status dictionary
process_status = {
	'slum': 0,
	'slum_unprocess': 0,
	
	'household': 0,
	'household_unprocess': 0,
	
	'proceess': 0,
	'success': 0,
	'fail': 0,
	
	'upload': 0
}

# get data from database
def fetch_db_records(query):
	global db_name
	global db_user
	global db_pwd
	global db_host
	global db_port
	
	conn = psycopg2.connect(database=db_name, user=db_user, password=db_pwd, host=db_host, port=db_port)
	#print("Opened database successfully")
	
	cur = conn.cursor()
	
	#print ("Query - ", query)
	cur.execute(query)
	
	#print ("Query executed Successfully")
	
	rows = cur.fetchall()
	#print ("Data fetched - Operation done successfully")
	#write_log("Data fetched : query  : " + query)
	
	conn.close()
	#print ("Closed database successfully")
	
	return rows;

#get single id list	
def get_list_ids(query):
	data = []
	db_row = fetch_db_records(query)
	
	for row in db_row:
		data.append(row[0])
	
	return data;

# get dictionary for question answer mapping
def get_question_answer(query):
	data = {}
	
	db_row = fetch_db_records(query)
	
	for row in db_row:
		key = row[0]
		val = row[1]
		
		if key not in data:
			data.setdefault(key, val)
		else:
			temp_val = data[key]
			if isinstance(temp_val, list):
				temp_val.append(val)
				data[key] = temp_val
			else:
				temp_lst = [temp_val, val]
				data[key] = temp_lst
		
	return data;

# get dictionary for question answer mapping of toilet
def get_toilet_question_answer(query):
	data = {}
	
	db_row = fetch_db_records(query)
	#print('data row -> ',db_row)
	for row in db_row:
		toilet = row[0]
		key = row[1]
		val = row[2]
		
		if toilet not in data:
			data.setdefault(toilet, {})
		
		toilet_dict = data[toilet]
		if key not in toilet_dict:
			toilet_dict.setdefault(key, val)
		else:
			temp_val = toilet_dict[key]
			
			if isinstance(temp_val, list):
				temp_val.append(val)
				toilet_dict[key] = temp_val
			else:
				temp_lst = [temp_val, val]
				toilet_dict[key] = temp_lst
	
	return data;

# get dictionary for question answer for per household
def get_household_wise_question_answer(query):
	data = {}
	
	db_row = fetch_db_records(query)
	#print('data row -> ',db_row)
	for row in db_row:
		household = row[0]
		key = row[1]
		val = row[2]
		
		if household not in data:
			data.setdefault(household, {})
		
		household_dict = data[household]
		if key not in household_dict:
			household_dict.setdefault(key, val)
		else:
			temp_val = household_dict[key]
			
			if isinstance(temp_val, list):
				temp_val.append(val)
				household_dict[key] = temp_val
			else:
				temp_lst = [temp_val, val]
				household_dict[key] = temp_lst

	return data;

# get dictionary for slum and its code
def get_slum_code(query):
	return get_question_answer(query)

# get household count
def get_household_count(query):
	data = 0
	db_row = fetch_db_records(query)
	
	for row in db_row:
		data = row[0]
		break;
	
	return data;

# get dictionary for household survey
def get_household_survey(query):
	data = {}
	db_row = fetch_db_records(query)
	
	for row in db_row:
		household = row[0]
		survey = row[1]
		
		if household not in data:
			data.setdefault(household, [])
			
		data[household].append(survey)
		
	return data;

# function to read excel file(xls)
def read_xml_excel(excelFile):	
	global option_dict
	global city_ward_slum_dict
	
	global question_map_dict
	
	#open excel file 
	#read question from sheet 1 into dict
	#map dict with option from sheet 2
	
	workbook = xlrd.open_workbook(excelFile)
	
	# List sheet names, and pull a sheet by name
	sheet_names = workbook.sheet_names()
	
	#print('Sheet Names', sheet_names)
	
	sheet_survey = workbook.sheet_by_index(0)
	sheet_choices = workbook.sheet_by_index(1)
	#sheet_settings = workbook.sheet_by_index(2)

	# read choice sheet to create option mapping dict 
	for row in range(sheet_choices.nrows):
		if row != 0:
			key = sheet_choices.cell_value(row, 0)
			value = sheet_choices.cell_value(row, 1)
			
			#print('row value ', name)
			if key not in option_dict:
				option_dict.setdefault(key, [])
				
			option_dict[key].append(value)
			
			city = sheet_choices.cell_value(row, 3)
			if city.strip():
				if city not in city_ward_slum_dict:
					city_ward_slum_dict.setdefault(city, {})
				
				if value not in city_ward_slum_dict[city]:
					city_ward_slum_dict[city].setdefault(value, [])
			
			admin_ward = sheet_choices.cell_value(row, 4)
			
			if admin_ward.strip():
				for city in city_ward_slum_dict:
					if admin_ward in city_ward_slum_dict[city]:
						city_ward_slum_dict[city][admin_ward].append(value)
						break
				
	
	#print("dict - ", option_dict)
	#print("city_ward_slum_dict - ", city_ward_slum_dict)
	
	# read choice sheet to create option mapping dict 
	for row in range(sheet_survey.nrows):
		if row != 0:
			name = sheet_survey.cell_value(row, 1)
			
			if name.strip() and not name.startswith("group_") and name not in question_map_dict:
				question_map_dict.setdefault(name, None)
	
	
	return;

# function to read excel file(xlsx)
def read_map_excel(excelFile):
	global question_map_dict
	global question_option_map_dict
	
	#open excel file 
	#read mapping for quetion and new xml key
	#create mapping dict
	
	workbook = openpyxl.load_workbook(excelFile)
	
	# List sheet names, and pull a sheet by name
	sheet_names = workbook.get_sheet_names()
	
	#print('Sheet Names', sheet_names)
	
	sheet_old_new_question_mapping = workbook.worksheets[0]
	sheet_old_new_option_mapping = workbook.worksheets[2]
	
	for row in sheet_old_new_question_mapping.iter_rows(row_offset=1):
		row_data  = []
		for cell in row:
			row_data.append(cell.value)
		
		question_id = row_data[5]
		if not (question_id is None):
			dict_key = row_data[1]
			
			question_map_dict[dict_key] = question_id
	
	#print("dict question map - ", question_map_dict)
	
	# set option mapping dict for all qustion mapped
	for row in sheet_old_new_option_mapping.iter_rows(row_offset=1):
		row_data  = []
		for cell in row:
			row_data.append(cell.value)
		
		question_id = row_data[8]
		
		if not (question_id is None):
			old_option = row_data[9]
			
			if not (old_option is None):
				new_option = row_data[2]
				
				if question_id not in question_option_map_dict:
					question_option_map_dict.setdefault(question_id, {})
				
				# in case multiple old option map to single option 
				if isinstance(old_option, int):
					question_option_map_dict[question_id].setdefault(old_option, new_option)
				else:
					old_option_list = old_option.split(',')
					for option_id in old_option_list:
						question_option_map_dict[question_id].setdefault(int(option_id), new_option)
	
	#print("dict optioin map - ", question_option_map_dict)
	
	return;

# get answer for xml key from question-answer dictionary
def get_answer(xml_key, fact_dict):
	global question_map_dict
	global question_option_map_dict
	answer = None
	
	if xml_key in question_map_dict:
		# get question id 
		fact_id = question_map_dict[xml_key]
		
		# if question not found then return
		if (fact_id is None):
			return None
		
		# get answer as return
		if fact_id in fact_dict:
			# get answer for fact 
			answer = fact_dict[fact_id]
			
			#check fact has options (single or multi select )
			if fact_id in question_option_map_dict:
				#check if answer is list - this is in case of multi select option
				if isinstance(answer, list):
					temp_answer = None
					for ans in answer:
						if ans and int(ans) in question_option_map_dict[fact_id]:
							temp_ans = question_option_map_dict[fact_id][int(ans)]
							if not (temp_ans is None):
								if temp_answer is None:
									temp_answer = str('')
								temp_answer = temp_answer + str(temp_ans) + ' '
					
					if not(temp_answer is None):
						answer = temp_answer.strip()
				else:
					if answer:
						answer_option = int(answer)
						if answer_option in question_option_map_dict[fact_id]:
							answer = question_option_map_dict[fact_id][answer_option]
		
	return answer;

def get_name_id(xml_key, answer_text):
	global option_dict
	
	answer = None
	
	#print('answer_text => ',answer_text)
	
	option_list = option_dict[xml_key]
	#print('option_list => ',option_list)
	
	if answer_text:
		name_list = answer_text.split(',')
		
		if not name_list:
			name_list = answer_text.split('/')
			
		#print('name_list => ',name_list)
	
		if name_list:
			for name in name_list:
				for option in option_list:
					#print(xml_key+"    "+name +"   "+option)
					if option.lower() == name.lower():
						answer = option
						
	
	return answer;

# get admin ward code for slum
def get_admin_ward(slum_code):
	admin_ward = None
	global city_ward_slum_dict
	
	for city, admin_ward_dict in iter(city_ward_slum_dict.items()):
		if admin_ward_dict:
			for ward, slum in iter(admin_ward_dict.items()):
				if slum_code in slum:
					admin_ward = ward
					break;
		if not (admin_ward is None):
			break;
	return admin_ward;

# get city id for admin ward
def get_city_id(admin_ward):
	city_id = None
	global city_ward_slum_dict
	
	for city, admin_ward_dict in iter(city_ward_slum_dict.items()):
		if admin_ward_dict:
			if admin_ward in admin_ward_dict:
				city_id = city
				break;
				
	return city_id;
	
	return;

# get formatted date in common format after parsing
def get_formatted_data(date_string):
	date_after_format = None
	date_converted = None
	
	date_format = ['%d/%m/%Y', '%d/%m/%y', '%d.%m.%Y', '%d.%m.%y', '%d-%m-%Y', '%d %b %Y', '%dth %b %Y', '%dth %B %Y', '%dth %B %Y.', '%dth %B, %Y', '%dth %b. %Y', 
					'%dst %b %Y', '%dnd %b %Y', '%drd %b %Y', '%dth %b %y', '%d/%b/%Y', '%dth %B %Y', '%dnd %B %Y', '%dst %B %Y', '%dnd %B %y', '%drd %B %Y', '%B %Y', 
					'%d/%m/%Y']
	
	for format in date_format:
		if date_converted is None:
			try:
				date_converted = datetime.datetime.strptime(date_string, format)
				break
			except:
				date_converted = None
				pass
	
	if date_converted:
		date_after_format = date_converted.strftime('%Y-%m-%d') #'%Y-%m-%dT%H:%M:%S-%z'
	
	return date_after_format;

# get area into converted square meter after parsing
def convert_area_from_square_meters(area_sq_m):
	area = 0
	
	if area_sq_m:
		area_sq_m = area_sq_m.lower()
		
		unformatted_area = area_sq_m.replace('sq.','').replace('m.','')
		
		unformatted_area =  unformatted_area.replace('m','')
		
		unformatted_area =  unformatted_area.replace('ts','')
		
		unformatted_area =  unformatted_area.replace('sq','')
		
		unformatted_area = unformatted_area.replace(' ','')
		unformatted_area = unformatted_area.replace(',','')
		
		try:
			#print('converted area=>'+unformatted_area+"  float => "+str(float(unformatted_area)))
			area = int(float(unformatted_area.strip()))
		except:
			write_log('unformatted area=>'+unformatted_area)
			#print('converted area=>'+unformatted_area)
			pass
			
		#print("final area => "+str(area))
	
	return area;

# get count of huts in slum (use for RA survey only)
def get_approximat_huts(slum_id, answer):
	huts_count = 0
	option = {1:75, 2:225, 3:400, 4:750, 5:1000}
	
	# get count of household in slum
	qry_ra_slum_household_count = "select count(id) from slum_data_household where slum_id=%s "
	
	huts_count = get_household_count(qry_ra_slum_household_count % slum_id)
	
	if huts_count == 0:
		if answer:
			huts_count = option[int(answer)]
			
	return huts_count;

# get area into converted square feet after parsing (use for RHS survey only)
def get_rhs_area_in_squar_feet(answer_sq_m):
	area = 0
	
	if answer_sq_m:
		try:
			area = int(answer_sq_m)
		except Exception:
			if '*' in answer_sq_m:
				area_size = answer_sq_m.split('*')
				area = int(area_size[0]) * int(area_size[1])
			elif ',' in answer_sq_m:
				area = int(answer_sq_m.replace(',',''))
			elif '/' in answer_sq_m:
				area = int(answer_sq_m.replace('/',''))
			elif isinstance(answer_sq_m, list):
				area = int(answer_sq_m[0])
			elif '`' in answer_sq_m:
				area = int(answer_sq_m.replace('`',''))
			elif 'sq' in answer_sq_m.lower():
				area_sq_m = answer_sq_m.lower()
				
				unformatted_area = area_sq_m.replace('sq','').replace('.ft','')
				
				area = int(unformatted_area.strip())
			else:
				raise Exception
			
	return area;

# get family member count after parsing (use for RHS survey only)
def get_rhs_family_member_count(answer_count):
	count = None
	
	if answer_count:
		try:
			count = int(answer_count)
		except Exception:
			if ',' in answer_count:
				count = int(answer_count.replace(',',''))
			elif '*' in answer_count:
				count = int(answer_count.replace('*',''))
			elif '?' in answer_count:
				count = int(answer_count.replace('?',''))
			elif '/' in answer_count:
				count = int(answer_count.replace('/',''))
			elif isinstance(answer_count, list):
				count = int(answer_count[0])
			else:
				raise Exception
				
	return count;

# get string from option list to store as text into xml
def get_option_text(option_list, answer):
	option_text = ''
	
	if answer:
		if isinstance(answer, list):
			temp_option = ''
			for id in answer:
				if id:
					temp_option += option_list[id] + ','
			
			option_text = temp_option[:-1]
		else:
			option_text = option_list[answer]
	
	return option_text;

# get xml string to store into xml file
def create_xml_string(xml_dict, repeat_dict, xml_root, xml_root_attr_id, xml_root_attr_version):
	xml_string = dicttoxml.dicttoxml(xml_dict, attr_type=False, custom_root=xml_root)
	#print(xml_string)
	#print("\n")
	root = ET.fromstring(xml_string)
	root.set('id', xml_root_attr_id)
	root.set('version', xml_root_attr_version)
	
	#repeat_dict = {'group_te3dx03' : { 'append_index' : 1, 'list' : toilet_info}}
	if repeat_dict:
		for key, val in repeat_dict.items():
			if val['list']:
				sub_ele = root.find(key)
				index = val['append_index']
				# create xml to be appened and append
				for sub_xml_dict in val['list']:
					sub_xml_string = dicttoxml.dicttoxml(sub_xml_dict, attr_type=False, root=False)
					#print('\n sub xml - %s -- '%index ,sub_xml_string)
					
					sub_root = ET.fromstring(sub_xml_string)
					
					sub_ele.insert(index, sub_root)
					index = index+1
	
	xml_string = ET.tostring(root, encoding="utf8", method='xml')
	#print('\n final xml -- ', xml_string)
	#write_log('created xml string to write')
	
	return root;

# create xml file on given location
def create_xml_file(xml_root, filename, folderpath):
	file = filename + ".xml"
	xml_file = os.path.join(folderpath, file)
	
	directory = os.path.dirname(xml_file)
	
	if not os.path.exists(directory):
		os.makedirs(directory)
	
	xml_tree = ET.ElementTree(xml_root)
	
	xml_tree.write(xml_file, xml_declaration=True, encoding='utf-8', method="xml")
	
	log_msg = "created xml file : " + xml_file
	#write_log(log_msg)
	#print(log_msg)
	
	return;

# write log into log file
def write_log(msg):
	global log_folder_path
	
	log_path = options_dict['log_folder_path'] if options_dict['log_folder_path'] else log_folder_path
	
	log_file = os.path.join(log_path, 'log_' + str(strftime("%d_%m_%Y", gmtime())) + '.txt')
	
	directory = os.path.dirname(log_file)
	
	if not os.path.exists(directory):
		os.makedirs(directory)
	
	cur_datetime = strftime("%Y-%m-%d %H:%M:%S", gmtime())
	
	msg_str = cur_datetime + "\t\t" + msg + "\n"
	
	filehandle = open(log_file, "a")
	
	filehandle.write(msg_str)
	
	filehandle.close()
	
	return;

# rest option dictionary use to set option/parameter use by method
def reset_survey_option():
	global options_dict
	
	options_dict['project'] = None
	
	options_dict['survey'] = None
	
	options_dict['mapped_excelFile'] = None
	
	options_dict['output_path'] = None
	
	options_dict['survey2'] = None
		
		
	return;

def set_survey_option(project, survey, mapped_excelFile, survey2=None):
	global options_dict
	
	options_dict['project'] = project
	
	options_dict['survey'] = survey
	
	options_dict['mapped_excelFile'] = mapped_excelFile
	
	options_dict['survey2'] = survey2
	
	return;

def set_survey_xml_option(root, root_attr_id, root_attr_version, formhub_uuid):
	global options_dict
	
	options_dict['xml_root'] = root
	options_dict['xml_root_attr_id'] = root_attr_id
	options_dict['xml_root_attr_version'] = root_attr_version
	options_dict['formhub_uuid'] = formhub_uuid
	
	return;

def set_survey_log_path_option(log_folder_path):
	global options_dict
	
	options_dict['log_folder_path'] = log_folder_path
	return;

def set_survey_output_path_option(output_path):
	global options_dict
	
	options_dict['output_path'] = output_path
	
	return;

def get_survey_option_output_path():
	global options_dict
	
	output_path = options_dict['output_path']
	
	return output_path;

# get photo - return photo name as answer and download photo from url (use for FF survey only)
def get_ff_photo(xml_key, fact_dict, download_folder_path):
	answer = None
	photo_name = None
	
	url_prefix = "http://survey.shelter-associates.org/media/"
	
	#check if answer is available
	if fact_dict:
		photo = get_answer(xml_key, fact_dict)
		
		# get photo path (relative path)
		photo_path = photo if not isinstance(photo, list) else photo[0]
		
		#\print('\nanswer => ', xml_key, ' => ', photo_path)
		
		if photo_path:
			if download_folder_path:
				photo_url = url_prefix + photo_path
				#print('photo_url => ', photo_url)
				
				photo_name = photo_path.split('/')[-1]
				#print('photo_name => ', photo_name)
				
				download_photo_path = os.path.join(download_folder_path, photo_name)
				#print('download_photo_path => ', download_photo_path)
				
				try:
					if not os.path.exists(download_folder_path):
						os.makedirs(download_folder_path)
				
					# Download the file from `url` and save it locally under `file_name`:
					response = requests.get(photo_url, stream=True, verify=False)
					with open(download_photo_path, 'wb') as out_file:
						shutil.copyfileobj(response.raw, out_file) # copy from temp location to final location
					del response
					
					# check if file is downloaded or not 
					if os.path.isfile(download_photo_path):
						answer = photo_name

				except Exception as ex:
					exception_log = 'Exception occurred for fetching photo \t  exception : '+ str(ex) +' \t  traceback : '+ traceback.format_exc()
					write_log(exception_log)
					pass
	
	return answer;


# display progress
def show_progress_bar (iteration, total_count, status_for = ''):
	prefix = status_for +' \t Progress status:'
	suffix = 'Completed'
	decimal_length = 0
	progress_char = '-'
	max_length = 50
	
	complete_percent = ("{0:." + str(decimal_length) + "f}").format(100 * (iteration / float(total_count)))
	
	progress_bar_length = int(max_length * iteration // total_count)
	
	progress_bar = progress_char * progress_bar_length + ' ' * (max_length - progress_bar_length)
	
	print('\r%s [%s] %s%% %s' % (prefix, progress_bar, complete_percent, suffix), end = '\r', flush=True)
    
	return;

# set count for slum process
def set_process_slum_count(total_slum, unprocess_slum):
	process_status['slum'] = total_slum
	process_status['slum_unprocess'] = unprocess_slum
	return;

# set count for household process	
def set_process_household_count(total_household, unprocess_household):
	process_status['household'] = total_household
	process_status['household_unprocess'] = unprocess_household
	return;

# set count for total records process
def set_process_count(total_process, success, fail):
	process_status['proceess'] = total_process
	process_status['success'] = success
	process_status['fail'] = fail
	return;

# set count for file upload
def set_upload_count(total_upload, success, fail):
	process_status['upload'] = total_upload
	process_status['success'] = success
	process_status['fail'] = fail
	return;

# display result after process
def show_process_status():
	if process_status['slum']:
		slum_status = 'Total slums : ' + str(process_status['slum'])
		slum_status += ('' if process_status['slum_unprocess'] == 0 else '\t total slums unable to process '+str(process_status['slum_unprocess']))
		
		print(slum_status)
		write_log(slum_status)
		
		process_status['slum'] = 0
		process_status['slum_unprocess'] = 0
	
	if process_status['household']:
		household_status = 'Total household in all slums : ' + str(process_status['household'])
		household_status += ('' if process_status['household_unprocess'] == 0 else '\t total household unable to process '+str(process_status['household_unprocess']))
		
		print(household_status)
		write_log(household_status)
		
		process_status['household'] = 0
		process_status['household_unprocess'] = 0
		
	
	if process_status['proceess']:
		create_xml_status = 'Total process records : ' + str(process_status['proceess'])
		create_xml_status += '\t Success : ' + str(process_status['success'])
		create_xml_status += '\t Fail : ' + str(process_status['fail'])
		
		print(create_xml_status)
		write_log(create_xml_status)
		
		process_status['proceess'] = 0
		process_status['success'] = 0
		process_status['fail'] = 0
		
	if process_status['upload']:
		upload_status = 'Total records for upload : ' + str(process_status['upload'])
		upload_status += '\t Success : ' + str(process_status['success'])
		upload_status += '\t Fail : ' + str(process_status['fail'])
		
		print(upload_status)
		write_log(upload_status)
		
		process_status['upload'] = 0
		process_status['success'] = 0
		process_status['fail'] = 0
	
	return;








